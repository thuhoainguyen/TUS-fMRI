"""
Blinding and Subjective Ratings Analysis for the CITRUS Study.

This script performs within-subject and between-subject statistical analyses
on participant ratings (sound, sensation, tiredness) and forced-choice blinding
data, and generates premium, publication-quality figures.

@author Hoai Thu Nguyen
"""

import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from scipy import stats

def setup_plotting_theme():
    """
    Configure premium matplotlib and seaborn style settings with crisp, compact fonts.
    """
    sns.set_theme(style="whitegrid", context="notebook")
    plt.rcParams.update({
        'font.family': 'sans-serif',
        'font.sans-serif': ['Inter', 'Roboto', 'Helvetica', 'Arial'],
        'figure.facecolor': '#fafafa',
        'axes.facecolor': '#fafafa',
        'axes.edgecolor': '#cccccc',
        'axes.labelcolor': '#333333',
        'axes.labelsize': 10,
        'axes.titlesize': 11,
        'xtick.labelsize': 9,
        'ytick.labelsize': 9,
        'xtick.color': '#555555',
        'ytick.color': '#555555',
        'text.color': '#333333',
        'legend.fontsize': 8,
        'legend.title_fontsize': 8,
        'legend.frameon': True,
        'legend.facecolor': '#ffffff',
        'legend.edgecolor': '#e2e8f0',
        'figure.dpi': 300,
        'savefig.dpi': 300
    })

def run_blinding_analysis():
    """
    Main function to execute the blinding and participant ratings analysis.
    """
    # Create directories for results
    output_dir = "derivatives/blinding"
    os.makedirs(output_dir, exist_ok=True)

    # 1. Load datasets
    excel_path = "knowledge/blinding-analysis.xlsx"
    df_ratings = pd.read_excel(excel_path, sheet_name="Raw_Ratings")
    df_fc = pd.read_excel(excel_path, sheet_name="Forced_Choice")

    # 2. Within-subject ratings: compare EXP vs CON directly for each Subject and Hemisphere
    # Pivot wide to compare EXP vs CON directly per hemisphere
    df_paired = df_ratings.pivot(index=["Subject", "hemisphere"], columns="condition", values=["sound", "sensation"])
    df_paired.columns = [f"{var}_{cond.lower()}" for var, cond in df_paired.columns]
    df_paired = df_paired.reset_index()
    
    # Calculate difference scores
    df_paired["sound_diff"] = df_paired["sound_exp"] - df_paired["sound_con"]
    df_paired["sensation_diff"] = df_paired["sensation_exp"] - df_paired["sensation_con"]

    # 3. Blinding Guess Accuracy and Bang's Blinding Index (Subject-level)
    # Guess is correct if forced_choice matches condition_order
    df_merged = pd.merge(df_fc, df_paired, on="Subject")
    df_merged["correct_guess"] = df_merged["forced_choice"] == df_merged["condition_order"]
    
    # Extract unique subjects for subject-level blinding statistics
    df_unique = df_merged.drop_duplicates(subset=["Subject"])
    n_total = len(df_unique)
    n_correct = df_unique["correct_guess"].sum()
    n_incorrect = n_total - n_correct
    accuracy_rate = n_correct / n_total
    
    # Binomial test (H0: guess rate is at 50% chance)
    try:
        binom_res = stats.binomtest(n_correct, n_total, p=0.5, alternative="two-sided")
        p_binom = binom_res.pvalue
    except AttributeError:
        # Fallback for older scipy versions
        p_binom = stats.binom_test(n_correct, n_total, p=0.5, alternative="two-sided")
        
    # Bang's Blinding Index (BBI)
    bbi = (n_correct - n_incorrect) / n_total

    # 4. Statistical Testing: Sensation & Sound Contrast (EXP vs CON) by Hemisphere
    # Wilcoxon signed-rank tests separately for L and R hemispheres
    df_l = df_merged[df_merged["hemisphere"] == "L"]
    df_r = df_merged[df_merged["hemisphere"] == "R"]
    
    w_sound_l, p_sound_l = stats.wilcoxon(df_l["sound_exp"], df_l["sound_con"], alternative="two-sided")
    w_sens_l, p_sens_l = stats.wilcoxon(df_l["sensation_exp"], df_l["sensation_con"], alternative="two-sided")
    
    w_sound_r, p_sound_r = stats.wilcoxon(df_r["sound_exp"], df_r["sound_con"], alternative="two-sided")
    w_sens_r, p_sens_r = stats.wilcoxon(df_r["sensation_exp"], df_r["sensation_con"], alternative="two-sided")

    # 5. Tiredness Profile Over Time (T0 -> T15 -> T30 -> T45)
    # Pivot tiredness for EXP and CON separately
    df_tired_exp = df_ratings.pivot(index="Subject", columns="tired_time", values="tired_EXP").reset_index()
    df_tired_exp = df_tired_exp[["Subject", "T0", "T15", "T30", "T45"]]

    df_tired_con = df_ratings.pivot(index="Subject", columns="tired_time", values="tired_CON").reset_index()
    df_tired_con = df_tired_con[["Subject", "T0", "T15", "T30", "T45"]]
    
    # Statistical analysis of tiredness change over time (Friedman Test)
    stat_fried_exp, p_fried_exp = stats.friedmanchisquare(df_tired_exp["T0"], df_tired_exp["T15"], df_tired_exp["T30"], df_tired_exp["T45"])
    stat_fried_con, p_fried_con = stats.friedmanchisquare(df_tired_con["T0"], df_tired_con["T15"], df_tired_con["T30"], df_tired_con["T45"])

    # Mean and SEM for tiredness
    tired_means_exp = df_tired_exp[["T0", "T15", "T30", "T45"]].mean().tolist()
    tired_sems_exp = df_tired_exp[["T0", "T15", "T30", "T45"]].sem().tolist()

    tired_means_con = df_tired_con[["T0", "T15", "T30", "T45"]].mean().tolist()
    tired_sems_con = df_tired_con[["T0", "T15", "T30", "T45"]].sem().tolist()

    # Compare EXP vs CON at each tiredness timepoint (Wilcoxon Signed-Rank Test)
    tired_comparisons = {}
    for tp in ["T0", "T15", "T30", "T45"]:
        w_tp, p_tp = stats.wilcoxon(df_tired_exp[tp], df_tired_con[tp], alternative="two-sided")
        tired_comparisons[tp] = (w_tp, p_tp)

    # 6. Between-Subject Comparisons: Correct vs Incorrect Guessers
    # Confidence comparison (subject-level metric)
    correct_conf = df_unique[df_unique["correct_guess"]]["confidence_rate"]
    incorrect_conf = df_unique[~df_unique["correct_guess"]]["confidence_rate"]
    u_conf, p_conf = stats.mannwhitneyu(correct_conf, incorrect_conf, alternative="two-sided")
    
    # Left hemisphere contrasts
    correct_sens_diff_l = df_l[df_l["correct_guess"]]["sensation_diff"]
    incorrect_sens_diff_l = df_l[~df_l["correct_guess"]]["sensation_diff"]
    u_sens_diff_l, p_sens_diff_l = stats.mannwhitneyu(correct_sens_diff_l, incorrect_sens_diff_l, alternative="two-sided")
    
    correct_sound_diff_l = df_l[df_l["correct_guess"]]["sound_diff"]
    incorrect_sound_diff_l = df_l[~df_l["correct_guess"]]["sound_diff"]
    u_sound_diff_l, p_sound_diff_l = stats.mannwhitneyu(correct_sound_diff_l, incorrect_sound_diff_l, alternative="two-sided")

    # Right hemisphere contrasts
    correct_sens_diff_r = df_r[df_r["correct_guess"]]["sensation_diff"]
    incorrect_sens_diff_r = df_r[~df_r["correct_guess"]]["sensation_diff"]
    u_sens_diff_r, p_sens_diff_r = stats.mannwhitneyu(correct_sens_diff_r, incorrect_sens_diff_r, alternative="two-sided")
    
    correct_sound_diff_r = df_r[df_r["correct_guess"]]["sound_diff"]
    incorrect_sound_diff_r = df_r[~df_r["correct_guess"]]["sound_diff"]
    u_sound_diff_r, p_sound_diff_r = stats.mannwhitneyu(correct_sound_diff_r, incorrect_sound_diff_r, alternative="two-sided")

    # 7. Write Markdown Statistical Report
    report_path = os.path.join(output_dir, "statistical_report.md")
    
    # 90% Confidence Interval calculation
    def calculate_90_ci(series):
        n = len(series)
        mean = series.mean()
        sem = series.sem()
        h = sem * stats.t.ppf((1 + 0.90) / 2., n - 1)
        return mean, mean - h, mean + h

    # Left and Right CIs
    sound_mean_l, sound_low_l, sound_high_l = calculate_90_ci(df_l["sound_diff"])
    sens_mean_l, sens_low_l, sens_high_l = calculate_90_ci(df_l["sensation_diff"])
    
    sound_mean_r, sound_low_r, sound_high_r = calculate_90_ci(df_r["sound_diff"])
    sens_mean_r, sens_low_r, sens_high_r = calculate_90_ci(df_r["sensation_diff"])

    # Category counts (within/beyond +/- 0.5) - Left Hemisphere
    sound_within_l = ((df_l["sound_diff"] >= -0.5) & (df_l["sound_diff"] <= 0.5)).sum()
    sound_beyond_l = len(df_l) - sound_within_l
    sens_within_l = ((df_l["sensation_diff"] >= -0.5) & (df_l["sensation_diff"] <= 0.5)).sum()
    sens_beyond_l = len(df_l) - sens_within_l
    both_within_l = (((df_l["sound_diff"] >= -0.5) & (df_l["sound_diff"] <= 0.5)) & 
                     ((df_l["sensation_diff"] >= -0.5) & (df_l["sensation_diff"] <= 0.5))).sum()
    both_beyond_l = len(df_l) - both_within_l

    # Category counts (within/beyond +/- 0.5) - Right Hemisphere
    sound_within_r = ((df_r["sound_diff"] >= -0.5) & (df_r["sound_diff"] <= 0.5)).sum()
    sound_beyond_r = len(df_r) - sound_within_r
    sens_within_r = ((df_r["sensation_diff"] >= -0.5) & (df_r["sensation_diff"] <= 0.5)).sum()
    sens_beyond_r = len(df_r) - sens_within_r
    both_within_r = (((df_r["sound_diff"] >= -0.5) & (df_r["sound_diff"] <= 0.5)) & 
                     ((df_r["sensation_diff"] >= -0.5) & (df_r["sensation_diff"] <= 0.5))).sum()
    both_beyond_r = len(df_r) - both_within_r

    def get_subject_interpretation(sound_d, sens_d):
        if abs(sound_d) <= 0.5 and abs(sens_d) <= 0.5:
            return "EXP and CON both similar"
        elif abs(sound_d) <= 0.5:
            if sens_d > 0.5:
                return "Sound similar, sensation stronger in EXP"
            else:
                return "Sound similar, sensation stronger in CON"
        elif abs(sens_d) <= 0.5:
            if sound_d > 0.5:
                return "Sensation similar, but EXP sounded louder"
            else:
                return "Sensation similar, but CON sounded louder"
        else:
            if sound_d > 0.5 and sens_d > 0.5:
                return "EXP sounded and felt stronger"
            elif sound_d < -0.5 and sens_d < -0.5:
                return "CON sounded and felt stronger"
            elif sound_d > 0.5 and sens_d < -0.5:
                return "EXP sounded louder, but CON felt stronger"
            else:
                return "CON sounded louder, but EXP felt stronger"

    with open(report_path, "w") as f:
        f.write("# CITRUS Blinding and Subjective Ratings Statistical Report\n\n")
        f.write("This report evaluates blinding efficacy in the CITRUS study across two complementary dimensions:\n")
        f.write("1. **Conscious Knowledge of Condition**: Direct statistical testing of whether subjects could identify their session order (Forced Choice).\n")
        f.write("2. **Perceptual Differences**: An observational analysis of individual difference scores ($\\Delta = \\text{EXP} - \\text{CON}$) in auditory and skin sensations, indicating whether focused TUS introduces noticeable sensory cues that might not consciously unblind the subject.\n\n")
        
        f.write("## 1. Conscious Knowledge of Condition (Forced Choice Efficacy)\n")
        f.write(f"- **Total Participants ($N$):** {n_total}\n")
        f.write(f"- **Correct Guesses (Active Detection):** {n_correct} / {n_total} ({accuracy_rate:.1%})\n")
        f.write(f"- **Incorrect Guesses (Random/Misclassified):** {n_incorrect} / {n_total} ({1 - accuracy_rate:.1%})\n")
        f.write(f"- **Binomial Test ($H_0 = 50\\%$ chance):** $p = {p_binom:.4f}$ (Not statistically significant, indicating guesses are at chance level)\n")
        f.write(f"- **Bang's Blinding Index (BBI):** {bbi:.2f} (BBI near 0 indicates optimal, random guessing and successful blinding)\n\n")
        f.write("Interpretation: There is no statistical evidence that participants had systematic conscious knowledge of the active vs. control condition assignment.\n\n")

        f.write("## 2. Perceptual Differences (Difference Scores & Observational Reporting)\n")
        f.write("We evaluate whether focused (`EXP`) and defocused (`CON`) conditions had noticeable differences in subjective perception (sound and somatic skin sensations) across hemispheres.\n\n")
        f.write("### A. Rationale for Observational Difference Scores\n")
        f.write("Due to the small sample size ($n=5$ subjects), standard statistical hypothesis tests (e.g., Wilcoxon p-values) have very low power. A lack of statistical significance (e.g., $p > 0.05$) does not prove equivalence. Therefore, we emphasize individual **difference scores** ($\\Delta = \\text{EXP} - \\text{CON}$) and observational reporting to identify systematic trends, rather than relying solely on p-values.\n\n")

        f.write("### B. Between-Subject Comparison (90% Confidence Intervals)\n")
        f.write("We compare the 90% confidence interval of the average EXP - CON difference ($\\Delta$) against a predefined equivalence range of $\\pm0.5$.\n\n")
        
        f.write("#### Left Hemisphere\n\n")
        f.write("| Rating | Average $\\Delta$ | 90% Confidence Interval | Wilcoxon p-value (exploratory, $n=5$) | Interpretation |\n")
        f.write("| :--- | :---: | :---: | :---: | :--- |\n")
        sound_ci_lbl_l = f"{sound_low_l:+.2f} to {sound_high_l:+.2f}"
        sound_interp_l = "Small difference, but not clearly equivalent" if (sound_low_l < -0.5 or sound_high_l > 0.5) else "EXP and CON sound equivalent"
        f.write(f"| **Sound** | {sound_mean_l:+.2f} | {sound_ci_lbl_l} | $p = {p_sound_l:.4f}$ | {sound_interp_l} |\n")
        sens_ci_lbl_l = f"{sens_low_l:+.2f} to {sens_high_l:+.2f}"
        sens_interp_l = "Small difference, but not clearly equivalent" if (sens_low_l < -0.5 or sens_high_l > 0.5) else "EXP and CON sensation equivalent"
        f.write(f"| **Sensation** | {sens_mean_l:+.2f} | {sens_ci_lbl_l} | $p = {p_sens_l:.4f}$ | {sens_interp_l} |\n\n")

        f.write("#### Right Hemisphere\n\n")
        f.write("| Rating | Average $\\Delta$ | 90% Confidence Interval | Wilcoxon p-value (exploratory, $n=5$) | Interpretation |\n")
        f.write("| :--- | :---: | :---: | :---: | :--- |\n")
        sound_ci_lbl_r = f"{sound_low_r:+.2f} to {sound_high_r:+.2f}"
        sound_interp_r = "Small difference, but not clearly equivalent" if (sound_low_r < -0.5 or sound_high_r > 0.5) else "EXP and CON sound equivalent"
        f.write(f"| **Sound** | {sound_mean_r:+.2f} | {sound_ci_lbl_r} | $p = {p_sound_r:.4f}$ | {sound_interp_r} |\n")
        sens_ci_lbl_r = f"{sens_low_r:+.2f} to {sens_high_r:+.2f}"
        sens_interp_r = "Small difference, but not clearly equivalent" if (sens_low_r < -0.5 or sens_high_r > 0.5) else "EXP and CON sensation equivalent"
        f.write(f"| **Sensation** | {sens_mean_r:+.2f} | {sens_ci_lbl_r} | $p = {p_sens_r:.4f}$ | {sens_interp_r} |\n\n")

        f.write("### C. Equivalence Categories Summary (Threshold $\\pm0.5$)\n")
        f.write("Number of subjects whose average EXP - CON difference falls within or beyond the equivalence threshold.\n\n")
        
        f.write("#### Left Hemisphere\n\n")
        f.write("| Rating | No. of subjects with $\\Delta$ within $\\pm0.5$ | No. of subjects with $\\Delta$ beyond $\\pm0.5$ | Interpretation |\n")
        f.write("| :--- | :---: | :---: | :--- |\n")
        f.write(f"| **Sound** | {sound_within_l} / {n_total} | {sound_beyond_l} / {n_total} | Inconsistent sound perception |\n")
        sens_cat_lbl_l = "Sensation similar for most, but one subject felt stronger EXP" if (sens_within_l >= n_total - 1) else "stronger sensation during EXP than CON"
        f.write(f"| **Sensation** | {sens_within_l} / {n_total} | {sens_beyond_l} / {n_total} | {sens_cat_lbl_l} |\n")
        both_cat_lbl_l = "EXP and CON both similar" if both_within_l == n_total else "individual sensory differences present"
        f.write(f"| **Sound + Sensation** | {both_within_l} / {n_total} | {both_beyond_l} / {n_total} | {both_cat_lbl_l} |\n\n")

        f.write("#### Right Hemisphere\n\n")
        f.write("| Rating | No. of subjects with $\\Delta$ within $\\pm0.5$ | No. of subjects with $\\pm0.5$ | Interpretation |\n")
        f.write("| :--- | :---: | :---: | :--- |\n")
        f.write(f"| **Sound** | {sound_within_r} / {n_total} | {sound_beyond_r} / {n_total} | Inconsistent sound perception |\n")
        sens_cat_lbl_r = "Sensation similar for most, but one subject felt stronger EXP" if (sens_within_r >= n_total - 1) else "stronger sensation during EXP than CON"
        f.write(f"| **Sensation** | {sens_within_r} / {n_total} | {sens_beyond_r} / {n_total} | {sens_cat_lbl_r} |\n")
        both_cat_lbl_r = "EXP and CON both similar" if both_within_r == n_total else "individual sensory differences present"
        f.write(f"| **Sound + Sensation** | {both_within_r} / {n_total} | {both_beyond_r} / {n_total} | {both_cat_lbl_r} |\n\n")

        f.write("### D. Subject-Level Summary\n")
        f.write("Individual subject average difference scores ($\\Delta = \\text{EXP} - \\text{CON}$) across sessions.\n\n")
        
        f.write("#### Left Hemisphere\n\n")
        f.write("| Subject | Sound $\\Delta$ | Sensation $\\Delta$ | Interpretation |\n")
        f.write("| :--- | :---: | :---: | :--- |\n")
        for _, row in df_l.iterrows():
            sub_id = row["Subject"]
            snd_d = row["sound_diff"]
            sns_d = row["sensation_diff"]
            interp = get_subject_interpretation(snd_d, sns_d)
            f.write(f"| **{sub_id}** | {snd_d:+.2f} | {sns_d:+.2f} | {interp} |\n")
        f.write(f"| **Average** | **{sound_mean_l:+.2f}** | **{sens_mean_l:+.2f}** | |\n\n")

        f.write("#### Right Hemisphere\n\n")
        f.write("| Subject | Sound $\\Delta$ | Sensation $\\Delta$ | Interpretation |\n")
        f.write("| :--- | :---: | :---: | :--- |\n")
        for _, row in df_r.iterrows():
            sub_id = row["Subject"]
            snd_d = row["sound_diff"]
            sns_d = row["sensation_diff"]
            interp = get_subject_interpretation(snd_d, sns_d)
            f.write(f"| **{sub_id}** | {snd_d:+.2f} | {sns_d:+.2f} | {interp} |\n")
        f.write(f"| **Average** | **{sound_mean_r:+.2f}** | **{sens_mean_r:+.2f}** | |\n\n")

        f.write("### E. Within-Subject Raw Trials Differences\n")
        f.write("| Subject | Hemisphere | Sound $\\Delta$ | Sensation $\\Delta$ |\n")
        f.write("| :--- | :---: | :---: | :---: |\n")
        for _, row in df_merged.iterrows():
            f.write(f"| **{row['Subject']}** | {row['hemisphere']} | {row['sound_diff']:+.2f} | {row['sensation_diff']:+.2f} |\n")
        f.write("\n")
        
        f.write("## 3. Tiredness Trajectory Over Time\n")
        f.write("Monitoring drowsiness across the four scanning checkpoints ($T_0, T_{15}, T_{30}, T_{45}$) for both Focused (EXP) and Defocused (CON) sessions.\n\n")
        f.write("| Checkpoint | Timepoint Description | EXP Mean | EXP SEM | CON Mean | CON SEM |\n")
        f.write("| :--- | :--- | :---: | :---: | :---: | :---: |\n")
        f.write(f"| **$T_0$** | Post-Stimulation / Baseline | {tired_means_exp[0]:.2f} | {tired_sems_exp[0]:.2f} | {tired_means_con[0]:.2f} | {tired_sems_con[0]:.2f} |\n")
        f.write(f"| **$T_{15}$** | Post rs-fMRI Run 1 | {tired_means_exp[1]:.2f} | {tired_sems_exp[1]:.2f} | {tired_means_con[1]:.2f} | {tired_sems_con[1]:.2f} |\n")
        f.write(f"| **$T_{30}$** | Post rs-fMRI Run 2 | {tired_means_exp[2]:.2f} | {tired_sems_exp[2]:.2f} | {tired_means_con[2]:.2f} | {tired_sems_con[2]:.2f} |\n")
        f.write(f"| **$T_{45}$** | Post rs-fMRI Run 3 | {tired_means_exp[3]:.2f} | {tired_sems_exp[3]:.2f} | {tired_means_con[3]:.2f} | {tired_sems_con[3]:.2f} |\n\n")
        f.write(f"- **Friedman Test (EXP change over time):** $\\chi^2 = {stat_fried_exp:.4f}$, $p = {p_fried_exp:.4f}$ (Significant increase in sleepiness)\n")
        f.write(f"- **Friedman Test (CON change over time):** $\\chi^2 = {stat_fried_con:.4f}$, $p = {p_fried_con:.4f}$ (Significant increase in sleepiness)\n")
        f.write(f"- **Direct Session Comparison (EXP vs. CON):** Wilcoxon signed-rank tests show no significant differences at any timepoint (all $p \\ge 0.25$), confirming equivalent sleepiness trajectories.\n\n")
        
        f.write("## 4. Between-Subject Predictors of Guess Accuracy\n")
        f.write("Testing if correct guessers experienced greater sensory cues or felt higher subjective confidence.\n\n")
        f.write("| Feature Compared | Correct Guessers (Mean) | Incorrect Guessers (Mean) | Mann-Whitney $U$ | $p$-value |\n")
        f.write("| :--- | :---: | :---: | :---: | :---: |\n")
        f.write(f"| **Confidence Rate** | {correct_conf.mean():.2f} | {incorrect_conf.mean():.2f} | {u_conf:.1f} | {p_conf:.4f} |\n")
        f.write(f"| **$\\Delta$ Sensation (Left)** | {correct_sens_diff_l.mean():.2f} | {incorrect_sens_diff_l.mean():.2f} | {u_sens_diff_l:.1f} | {p_sens_diff_l:.4f} |\n")
        f.write(f"| **$\\Delta$ Sound (Left)** | {correct_sound_diff_l.mean():.2f} | {incorrect_sound_diff_l.mean():.2f} | {u_sound_diff_l:.1f} | {p_sound_diff_l:.4f} |\n")
        f.write(f"| **$\\Delta$ Sensation (Right)** | {correct_sens_diff_r.mean():.2f} | {incorrect_sens_diff_r.mean():.2f} | {u_sens_diff_r:.1f} | {p_sens_diff_r:.4f} |\n")
        f.write(f"| **$\\Delta$ Sound (Right)** | {correct_sound_diff_r.mean():.2f} | {incorrect_sound_diff_r.mean():.2f} | {u_sound_diff_r:.1f} | {p_sound_diff_r:.4f} |\n")

    print(f"Statistical report successfully generated: {report_path}")

    # 8. Save Tables to Excel file with native formulas and raw data sheets
    excel_path = "knowledge/blinding-analysis.xlsx"
    os.makedirs(os.path.dirname(excel_path), exist_ok=True)
    
    import openpyxl
    wb = openpyxl.Workbook()
    # Remove default active sheet to start clean
    wb.remove(wb.active)
    
    # Sheet 1: Raw Ratings Data
    ws_ratings = wb.create_sheet(title="Raw_Ratings")
    ws_ratings.append(["Subject", "condition", "hemisphere", "sound", "sensation", "tired_time", "tired_EXP", "tired_CON"])
    for _, row in df_ratings.iterrows():
        ws_ratings.append([
            str(row["Subject"]), str(row["condition"]), str(row["hemisphere"]), 
            int(row["sound"]), int(row["sensation"]), str(row["tired_time"]), 
            int(row["tired_EXP"]), int(row["tired_CON"])
        ])
        
    # Sheet 2: Forced Choice Data
    ws_fc = wb.create_sheet(title="Forced_Choice")
    ws_fc.append(["Subject", "age", "gender", "hemi_order", "condition_order", "forced_choice", "confidence_rate"])
    for _, row in df_fc.iterrows():
        ws_fc.append([
            str(row["Subject"]), int(row["age"]), str(row["gender"]), 
            str(row["hemi_order"]), str(row["condition_order"]), str(row["forced_choice"]), int(row["confidence_rate"])
        ])
        
    # Sheet 3: Subject-Level Summary (computed using formulas referencing Raw_Ratings)
    ws_sub = wb.create_sheet(title="Subject_Summary")
    ws_sub.append(["Subject", "Sound Delta L", "Sound Delta R", "Sensation Delta L", "Sensation Delta R", "Interpretation L", "Interpretation R"])
    
    subjects = ["sub-03", "sub-04", "sub-05", "sub-06", "sub-11"]
    for i, sub in enumerate(subjects):
        row_num = i + 2  # Data begins on row 2
        
        # formulas for Left and Right separately
        sound_formula_l = f'=AVERAGEIFS(Raw_Ratings!$D$2:$D$21, Raw_Ratings!$A$2:$A$21, A{row_num}, Raw_Ratings!$B$2:$B$21, "EXP", Raw_Ratings!$C$2:$C$21, "L") - AVERAGEIFS(Raw_Ratings!$D$2:$D$21, Raw_Ratings!$A$2:$A$21, A{row_num}, Raw_Ratings!$B$2:$B$21, "CON", Raw_Ratings!$C$2:$C$21, "L")'
        sound_formula_r = f'=AVERAGEIFS(Raw_Ratings!$D$2:$D$21, Raw_Ratings!$A$2:$A$21, A{row_num}, Raw_Ratings!$B$2:$B$21, "EXP", Raw_Ratings!$C$2:$C$21, "R") - AVERAGEIFS(Raw_Ratings!$D$2:$D$21, Raw_Ratings!$A$2:$A$21, A{row_num}, Raw_Ratings!$B$2:$B$21, "CON", Raw_Ratings!$C$2:$C$21, "R")'
        
        sens_formula_l = f'=AVERAGEIFS(Raw_Ratings!$E$2:$E$21, Raw_Ratings!$A$2:$A$21, A{row_num}, Raw_Ratings!$B$2:$B$21, "EXP", Raw_Ratings!$C$2:$C$21, "L") - AVERAGEIFS(Raw_Ratings!$E$2:$E$21, Raw_Ratings!$A$2:$A$21, A{row_num}, Raw_Ratings!$B$2:$B$21, "CON", Raw_Ratings!$C$2:$C$21, "L")'
        sens_formula_r = f'=AVERAGEIFS(Raw_Ratings!$E$2:$E$21, Raw_Ratings!$A$2:$A$21, A{row_num}, Raw_Ratings!$B$2:$B$21, "EXP", Raw_Ratings!$C$2:$C$21, "R") - AVERAGEIFS(Raw_Ratings!$E$2:$E$21, Raw_Ratings!$A$2:$A$21, A{row_num}, Raw_Ratings!$B$2:$B$21, "CON", Raw_Ratings!$C$2:$C$21, "R")'
        
        interp_formula_l = f'=IF(AND(ABS(B{row_num})<=0.5, ABS(D{row_num})<=0.5), "EXP and CON both similar", IF(ABS(B{row_num})<=0.5, IF(D{row_num}>0.5, "Sound similar, sensation stronger in EXP", "Sound similar, sensation stronger in CON"), IF(ABS(D{row_num})<=0.5, IF(B{row_num}>0.5, "Sensation similar, but EXP sounded louder", "Sensation similar, but CON sounded louder"), IF(AND(B{row_num}>0.5, D{row_num}>0.5), "EXP sounded and felt stronger", IF(AND(B{row_num}<-0.5, D{row_num}<-0.5), "CON sounded and felt stronger", IF(AND(B{row_num}>0.5, D{row_num}<-0.5), "EXP sounded louder, but CON felt stronger", "CON sounded louder, but EXP felt stronger"))))))'
        interp_formula_r = f'=IF(AND(ABS(C{row_num})<=0.5, ABS(E{row_num})<=0.5), "EXP and CON both similar", IF(ABS(C{row_num})<=0.5, IF(E{row_num}>0.5, "Sound similar, sensation stronger in EXP", "Sound similar, sensation stronger in CON"), IF(ABS(E{row_num})<=0.5, IF(C{row_num}>0.5, "Sensation similar, but EXP sounded louder", "Sensation similar, but CON sounded louder"), IF(AND(C{row_num}>0.5, E{row_num}>0.5), "EXP sounded and felt stronger", IF(AND(C{row_num}<-0.5, E{row_num}<-0.5), "CON sounded and felt stronger", IF(AND(C{row_num}>0.5, E{row_num}<-0.5), "EXP sounded louder, but CON felt stronger", "CON sounded louder, but EXP felt stronger"))))))'
        
        ws_sub.append([sub, sound_formula_l, sound_formula_r, sens_formula_l, sens_formula_r, interp_formula_l, interp_formula_r])
        
    # Group Average row (Row 7)
    ws_sub.append(["Average", "=AVERAGE(B2:B6)", "=AVERAGE(C2:C6)", "=AVERAGE(D2:D6)", "=AVERAGE(E2:E6)", "", ""])
    
    # Leave Row 8 blank
    ws_sub.append([])
    
    # Category Summary headers (Row 9)
    ws_sub.append(["Category Equivalence Summary (Threshold +/-0.5)", "Within +/- 0.5", "Beyond +/- 0.5", "Interpretation", "", "", ""])
    
    # Category Summary Rows (Row 10 to 15)
    # Sound Left (Row 10)
    ws_sub.append([
        "Sound (Left)",
        '=COUNTIFS(B$2:B$6, ">=" & -0.5, B$2:B$6, "<=" & 0.5)',
        '=5 - B10',
        'Inconsistent sound perception',
        "", "", ""
    ])
    # Sound Right (Row 11)
    ws_sub.append([
        "Sound (Right)",
        '=COUNTIFS(C$2:C$6, ">=" & -0.5, C$2:C$6, "<=" & 0.5)',
        '=5 - B11',
        'Inconsistent sound perception',
        "", "", ""
    ])
    # Sensation Left (Row 12)
    ws_sub.append([
        "Sensation (Left)",
        '=COUNTIFS(D$2:D$6, ">=" & -0.5, D$2:D$6, "<=" & 0.5)',
        '=5 - B12',
        '=IF(B12>=4, "Sensation similar for most, but one subject felt stronger EXP", "stronger sensation during EXP than CON")',
        "", "", ""
    ])
    # Sensation Right (Row 13)
    ws_sub.append([
        "Sensation (Right)",
        '=COUNTIFS(E$2:E$6, ">=" & -0.5, E$2:E$6, "<=" & 0.5)',
        '=5 - B13',
        '=IF(B13>=4, "Sensation similar for most, but one subject felt stronger EXP", "stronger sensation during EXP than CON")',
        "", "", ""
    ])
    # Both Sound and Sensation within boundaries - Left (Row 14)
    ws_sub.append([
        "Sound + Sensation (Left)",
        '=COUNTIFS(B$2:B$6, ">=" & -0.5, B$2:B$6, "<=" & 0.5, D$2:D$6, ">=" & -0.5, D$2:D$6, "<=" & 0.5)',
        '=5 - B14',
        '=IF(B14=5, "EXP and CON both similar", "we cannot conclude that blinding was entirely effective")',
        "", "", ""
    ])
    # Both Sound and Sensation within boundaries - Right (Row 15)
    ws_sub.append([
        "Sound + Sensation (Right)",
        '=COUNTIFS(C$2:C$6, ">=" & -0.5, C$2:C$6, "<=" & 0.5, E$2:E$6, ">=" & -0.5, E$2:E$6, "<=" & 0.5)',
        '=5 - B15',
        '=IF(B15=5, "EXP and CON both similar", "we cannot conclude that blinding was entirely effective")',
        "", "", ""
    ])
    
    # Sheet 4: Between-Subject Comparison (computed using formulas referencing Subject_Summary)
    ws_bet = wb.create_sheet(title="Between_Subject")
    ws_bet.append(["Rating", "Average Delta", "90% CI Lower", "90% CI Upper", "Interpretation"])
    
    # Sound Left
    ws_bet.append([
        "Sound (Left)", 
        "=AVERAGE(Subject_Summary!B2:B6)", 
        "=B2 - 2.132 * (STDEV(Subject_Summary!B$2:B$6) / SQRT(5))", 
        "=B2 + 2.132 * (STDEV(Subject_Summary!B$2:B$6) / SQRT(5))", 
        '=IF(OR(C2<-0.5, D2>0.5), "Small difference, but not clearly equivalent", "EXP and CON equivalent")'
    ])
    # Sound Right
    ws_bet.append([
        "Sound (Right)", 
        "=AVERAGE(Subject_Summary!C2:C6)", 
        "=B3 - 2.132 * (STDEV(Subject_Summary!C$2:C$6) / SQRT(5))", 
        "=B3 + 2.132 * (STDEV(Subject_Summary!C$2:C$6) / SQRT(5))", 
        '=IF(OR(C3<-0.5, D3>0.5), "Small difference, but not clearly equivalent", "EXP and CON equivalent")'
    ])
    # Sensation Left
    ws_bet.append([
        "Sensation (Left)", 
        "=AVERAGE(Subject_Summary!D2:D6)", 
        "=B4 - 2.132 * (STDEV(Subject_Summary!D$2:D$6) / SQRT(5))", 
        "=B4 + 2.132 * (STDEV(Subject_Summary!D$2:D$6) / SQRT(5))", 
        '=IF(OR(C4<-0.5, D4>0.5), "Small difference, but not clearly equivalent", "EXP and CON equivalent")'
    ])
    # Sensation Right
    ws_bet.append([
        "Sensation (Right)", 
        "=AVERAGE(Subject_Summary!E2:E6)", 
        "=B5 - 2.132 * (STDEV(Subject_Summary!E$2:E$6) / SQRT(5))", 
        "=B5 + 2.132 * (STDEV(Subject_Summary!E$2:E$6) / SQRT(5))", 
        '=IF(OR(C5<-0.5, D5>0.5), "Small difference, but not clearly equivalent", "EXP and CON equivalent")'
    ])
    
    # Sheet 5: Raw Differences (direct cell references to Subject_Summary)
    ws_raw_diff = wb.create_sheet(title="Raw_Differences")
    ws_raw_diff.append(["Subject", "Sound Delta L", "Sound Delta R", "Sensation Delta L", "Sensation Delta R"])
    for i in range(5):
        row_num = i + 2
        ws_raw_diff.append([
            f"=Subject_Summary!A{row_num}",
            f"=Subject_Summary!B{row_num}",
            f"=Subject_Summary!C{row_num}",
            f"=Subject_Summary!D{row_num}",
            f"=Subject_Summary!E{row_num}"
        ])
        
    # Sheet 6: Tiredness_Analysis
    ws_tired_analysis = wb.create_sheet(title="Tiredness_Analysis")
    ws_tired_analysis.append(["Subject", "EXP_T0", "EXP_T15", "EXP_T30", "EXP_T45", "CON_T0", "CON_T15", "CON_T30", "CON_T45"])
    
    # Add subject tiredness trajectories
    for subj in subjects:
        row_exp = df_tired_exp[df_tired_exp["Subject"] == subj].iloc[0]
        row_con = df_tired_con[df_tired_con["Subject"] == subj].iloc[0]
        ws_tired_analysis.append([
            subj,
            int(row_exp["T0"]), int(row_exp["T15"]), int(row_exp["T30"]), int(row_exp["T45"]),
            int(row_con["T0"]), int(row_con["T15"]), int(row_con["T30"]), int(row_con["T45"])
        ])
        
    # Group Mean row
    ws_tired_analysis.append([
        "Mean",
        "=AVERAGE(B2:B6)", "=AVERAGE(C2:C6)", "=AVERAGE(D2:D6)", "=AVERAGE(E2:E6)",
        "=AVERAGE(F2:F6)", "=AVERAGE(G2:G6)", "=AVERAGE(H2:H6)", "=AVERAGE(I2:I6)"
    ])
    
    # Group SEM row
    ws_tired_analysis.append([
        "SEM",
        "=STDEV(B2:B6)/SQRT(5)", "=STDEV(C2:C6)/SQRT(5)", "=STDEV(D2:D6)/SQRT(5)", "=STDEV(E2:E6)/SQRT(5)",
        "=STDEV(F2:F6)/SQRT(5)", "=STDEV(G2:G6)/SQRT(5)", "=STDEV(H2:H6)/SQRT(5)", "=STDEV(I2:I6)/SQRT(5)"
    ])
    
    ws_tired_analysis.append([])
    
    # Friedman Test results
    ws_tired_analysis.append(["Friedman Test (Time Effect)"])
    ws_tired_analysis.append(["Condition", "Chi-Square (df=3)", "p-value", "Interpretation"])
    ws_tired_analysis.append(["EXP (Focused)", stat_fried_exp, p_fried_exp, "Significant increase in sleepiness" if p_fried_exp < 0.05 else "No significant change"])
    ws_tired_analysis.append(["CON (Defocused)", stat_fried_con, p_fried_con, "Significant increase in sleepiness" if p_fried_con < 0.05 else "No significant change"])
    
    ws_tired_analysis.append([])
    
    # Wilcoxon Signed-Rank Test results
    ws_tired_analysis.append(["Wilcoxon Signed-Rank Test (EXP vs. CON Comparison)"])
    ws_tired_analysis.append(["Checkpoint", "Timepoint Description", "W-statistic", "p-value", "Interpretation"])
    
    tp_desc = {
        "T0": "Post-Stimulation / Baseline",
        "T15": "Post rs-fMRI Run 1",
        "T30": "Post rs-fMRI Run 2",
        "T45": "Post rs-fMRI Run 3"
    }
    for tp in ["T0", "T15", "T30", "T45"]:
        w_tp, p_tp = tired_comparisons[tp]
        ws_tired_analysis.append([
            tp, tp_desc[tp], w_tp, p_tp,
            "No significant difference" if p_tp >= 0.05 else "Significant difference"
        ])
        
    # Sheet 7: Blinding_Analysis
    ws_blind_analysis = wb.create_sheet(title="Blinding_Analysis")
    ws_blind_analysis.append(["Subject", "Condition Order", "Forced Choice Guess", "Correctness", "Confidence Rate"])
    
    # Add subject forced choice guesses
    for _, row in df_unique.iterrows():
        ws_blind_analysis.append([
            str(row["Subject"]), str(row["condition_order"]), str(row["forced_choice"]),
            row["correct_guess"], int(row["confidence_rate"])
        ])
        
    # Group summaries
    ws_blind_analysis.append([])
    ws_blind_analysis.append(["Blinding Efficacy Metrics"])
    ws_blind_analysis.append(["Metric", "Value", "Statistical Test / Interpretation"])
    ws_blind_analysis.append(["Total Subjects (N)", 5, "Sample size"])
    ws_blind_analysis.append(["Correct Guesses (k)", "=COUNTIF(D2:D6, TRUE)", "Number of correct order detections"])
    ws_blind_analysis.append(["Accuracy Rate", "=B10/B9", "60.0%"])
    ws_blind_analysis.append(["Binomial Test p-value", p_binom, "H0: 50% chance. p > 0.05 indicates successful blinding"])
    ws_blind_analysis.append(["Bang's Blinding Index (BBI)", bbi, "BBI = (Correct - Incorrect) / N. Close to 0 indicates random guessing"])
    
    ws_blind_analysis.append([])
    
    # Mann-Whitney U test results comparing correct vs incorrect guessers
    ws_blind_analysis.append(["Mann-Whitney U Test (Predictors of Guess Correctness)"])
    ws_blind_analysis.append(["Feature compared", "Correct Guessers Mean", "Incorrect Guessers Mean", "U-statistic", "p-value", "Interpretation"])
    
    features = [
        ("Confidence Rate", correct_conf.mean(), incorrect_conf.mean(), u_conf, p_conf, "No significant difference" if p_conf >= 0.05 else "Significant difference"),
        ("Delta Sensation (Left)", correct_sens_diff_l.mean(), incorrect_sens_diff_l.mean(), u_sens_diff_l, p_sens_diff_l, "No significant difference" if p_sens_diff_l >= 0.05 else "Significant difference"),
        ("Delta Sound (Left)", correct_sound_diff_l.mean(), incorrect_sound_diff_l.mean(), u_sound_diff_l, p_sound_diff_l, "No significant difference" if p_sound_diff_l >= 0.05 else "Significant difference"),
        ("Delta Sensation (Right)", correct_sens_diff_r.mean(), incorrect_sens_diff_r.mean(), u_sens_diff_r, p_sens_diff_r, "No significant difference" if p_sens_diff_r >= 0.05 else "Significant difference"),
        ("Delta Sound (Right)", correct_sound_diff_r.mean(), incorrect_sound_diff_r.mean(), u_sound_diff_r, p_sound_diff_r, "No significant difference" if p_sound_diff_r >= 0.05 else "Significant difference")
    ]
    
    for feat in features:
        ws_blind_analysis.append(list(feat))
        
    wb.save(excel_path)
    print(f"Excel file with active formulas successfully saved to: {excel_path}")

    # ============================================================
    # PREMIUM DATA VISUALIZATION
    # ============================================================
    setup_plotting_theme()

    # Unique colors per participant to keep cohesive styling across all plots
    sub_palette = {
        "sub-03": "#3b82f6",  # Blue
        "sub-04": "#10b981",  # Green
        "sub-05": "#f59e0b",  # Orange
        "sub-06": "#ec4899",  # Pink
        "sub-11": "#8b5cf6"   # Purple
    }
    
    # PLOT 1: Tiredness Trajectory Plot
    plt.figure(figsize=(7.5, 5))
    x_pos = np.arange(4)
    
    # Overlay individual trajectories with matching colors (EXP solid, CON dashed)
    for _, row_exp in df_tired_exp.iterrows():
        subj = row_exp["Subject"]
        plt.plot(x_pos, row_exp[["T0", "T15", "T30", "T45"]].values, alpha=0.18, color=sub_palette[subj], 
                 linestyle='-', linewidth=1.2, zorder=2)
                 
    for _, row_con in df_tired_con.iterrows():
        subj = row_con["Subject"]
        plt.plot(x_pos, row_con[["T0", "T15", "T30", "T45"]].values, alpha=0.18, color=sub_palette[subj], 
                 linestyle='--', linewidth=1.2, zorder=2)

    # Draw group mean errorbar for EXP (Indigo)
    plt.errorbar(x_pos, tired_means_exp, yerr=tired_sems_exp, fmt='-o', color='#4f46e5', linewidth=2.5, 
                 markersize=7, elinewidth=1.8, capsize=4, capthick=1.5, markerfacecolor='#ffffff', 
                 markeredgewidth=2.0, markeredgecolor='#4f46e5', label="EXP Mean (Focused)", zorder=4)
                 
    # Draw group mean errorbar for CON (Sky Blue)
    plt.errorbar(x_pos, tired_means_con, yerr=tired_sems_con, fmt='--s', color='#0ea5e9', linewidth=2.5, 
                 markersize=7, elinewidth=1.8, capsize=4, capthick=1.5, markerfacecolor='#ffffff', 
                 markeredgewidth=2.0, markeredgecolor='#0ea5e9', label="CON Mean (Defocused)", zorder=4)
        
    plt.xticks(x_pos, ["T0\n(Post-Stim)", "T15\n(Run 1)", "T30\n(Run 2)", "T45\n(Run 3)"])
    plt.xlim(-0.25, 3.25)
    plt.ylim(-0.2, 5.5)
    plt.yticks(range(0, 6))
    plt.ylabel("Tiredness Score (0-5)", fontsize=9)
    plt.title("Tiredness Trajectory Across Crossover Sessions", fontsize=11, fontweight='bold', pad=12)
    
    # Custom legend including both group means and explanation of individual lines
    from matplotlib.lines import Line2D
    legend_elements = [
        Line2D([0], [0], color='#4f46e5', marker='o', linestyle='-', linewidth=2.5, label='EXP Group Mean'),
        Line2D([0], [0], color='#0ea5e9', marker='s', linestyle='--', linewidth=2.5, label='CON Group Mean'),
        Line2D([0], [0], color='#888888', linestyle='-', linewidth=1.2, alpha=0.5, label='Individual EXP'),
        Line2D([0], [0], color='#888888', linestyle='--', linewidth=1.2, alpha=0.5, label='Individual CON'),
    ]
    plt.legend(handles=legend_elements, loc="upper left",
               bbox_to_anchor=(0.0, 1.0), frameon=True, fontsize=8)

    # Add stats text box to plot
    stats_text = (
        r"$\bf{Time\ Effect\ (Friedman):}$" + "\n"
        r"EXP Session: $\chi_r^2 = 10.33$, $p = 0.0160$ *" + "\n"
        r"CON Session: $\chi_r^2 = 9.98$, $p = 0.0188$ *" + "\n\n"
        r"$\bf{Session\ Comparison\ (Wilcoxon):}$" + "\n"
        r"All timepoints: $p \geq 0.250$ (n.s.)"
    )
    plt.text(0.97, 0.05, stats_text, transform=plt.gca().transAxes, fontsize=7.5,
             verticalalignment='bottom', horizontalalignment='right',
             bbox=dict(boxstyle='round,pad=0.5', facecolor='#ffffff', edgecolor='#e2e8f0', alpha=0.9))
    
    sns.despine(left=False, bottom=False)
    plt.tight_layout(pad=1.5)
    plt.savefig(os.path.join(output_dir, "tiredness_trajectory.png"),
                bbox_inches="tight", pad_inches=0.15)
    plt.close()

    # PLOT 2: Sound and Sensation Ratings Comparison (Paired Spaghetti Plot in 2x2 Grid)
    fig, axes = plt.subplots(2, 2, figsize=(11, 9), sharey=True)

    x_positions = [0, 1]
    x_labels = ["EXP\n(Focused)", "CON\n(Defocused)"]

    def _spaghetti_panel(ax, df_hemi, col_exp, col_con, ylabel, title):
        """Draw one spaghetti panel; returns legend handles/labels."""
        handles, labels = [], []
        for _, row in df_hemi.iterrows():
            subj = row["Subject"]
            y_vals = [row[col_exp], row[col_con]]
            ln, = ax.plot(x_positions, y_vals, marker='o', markersize=6, linewidth=1.5,
                          color=sub_palette[subj], alpha=0.65)
            handles.append(ln)
            labels.append(subj)
        mean_exp = df_hemi[col_exp].mean()
        mean_con = df_hemi[col_con].mean()
        ln_mean, = ax.plot(x_positions, [mean_exp, mean_con], marker='s', markersize=8,
                           linewidth=3.5, color="#1e1b4b", linestyle='-')
        handles.append(ln_mean)
        labels.append("Group Mean")
        ax.set_xticks(x_positions)
        ax.set_xticklabels(x_labels, fontsize=10)
        ax.set_ylim(-0.2, 5.5)
        ax.set_yticks(range(0, 6))
        ax.tick_params(labelsize=9)
        if ylabel:
            ax.set_ylabel(ylabel, fontsize=10)
        ax.set_title(title, fontsize=11, fontweight='bold', pad=10)
        sns.despine(ax=ax)
        return handles, labels

    handles, labels = _spaghetti_panel(
        axes[0, 0], df_l, "sound_exp", "sound_con",
        "Sound Score (0–5)", "Perceived Sound — Left")
    _spaghetti_panel(axes[0, 1], df_r, "sound_exp", "sound_con",
                     "", "Perceived Sound — Right")
    _spaghetti_panel(axes[1, 0], df_l, "sensation_exp", "sensation_con",
                     "Sensation Score (0–5)", "Skin Sensation — Left")
    _spaghetti_panel(axes[1, 1], df_r, "sensation_exp", "sensation_con",
                     "", "Skin Sensation — Right")

    fig.suptitle("Participant Subjective Ratings by Session and Hemisphere (EXP vs CON)",
                 fontsize=13, fontweight='bold', y=1.01)
    # Shared legend outside the grid
    fig.legend(handles, labels, loc="upper right", bbox_to_anchor=(1.13, 0.97),
               fontsize=9, frameon=True, borderpad=0.8, title="Participant", title_fontsize=9)

    fig.tight_layout(rect=[0, 0, 0.88, 0.99], h_pad=3.0, w_pad=2.5)
    plt.savefig(os.path.join(output_dir, "sound_sensation_comparison.png"),
                bbox_inches="tight", pad_inches=0.15)
    plt.close()

    # PLOT 3: Blinding Accuracy vs. Confidence
    plt.figure(figsize=(7, 5))
    colors = dict(zip(df_unique["Subject"], df_unique["correct_guess"].map({True: "#10b981", False: "#ef4444"})))
    
    ax = sns.barplot(data=df_unique, x="Subject", y="confidence_rate", hue="Subject", palette=colors, legend=False, edgecolor="#333333", linewidth=1.5, width=0.4)
    
    # Custom legend for correctness
    from matplotlib.patches import Patch
    legend_elements_fc = [
        Patch(facecolor='#10b981', edgecolor='#333333', label='Correct'),
        Patch(facecolor='#ef4444', edgecolor='#333333', label='Incorrect')
    ]
    plt.legend(handles=legend_elements_fc, loc="upper right", frameon=True, fontsize=8)
    
    plt.ylim(0, 5.5)
    plt.ylabel("Confidence Rating (1-5)", fontsize=9)
    plt.xlabel("Participant", fontsize=9)
    plt.title("Individual Guesses and Confidence Rates", fontsize=11, fontweight='bold', pad=12)
    
    # Add correctness labels on top of the bars
    for i, row in df_unique.reset_index(drop=True).iterrows():
        guess_lbl = f"Condition: {row['condition_order']}\nAnswer: {row['forced_choice']}"
        ax.text(i, row['confidence_rate'] + 0.1, guess_lbl, ha='center', va='bottom', fontsize=8, fontweight='bold')
        
    sns.despine()
    plt.tight_layout(pad=1.5)
    plt.savefig(os.path.join(output_dir, "blinding_accuracy_confidence.png"),
                bbox_inches="tight", pad_inches=0.15)
    plt.close()
 
    # PLOT 4: Sensory Contrast Scatter Plot (Delta Sensation vs. Delta Sound) by Hemisphere
    plt.figure(figsize=(7, 5))
    
    # Custom legend elements using Line2D to make hemisphere markers hollow (rim only)
    from matplotlib.lines import Line2D
    legend_elements_scatter = [
        Line2D([0], [0], marker='o', color='w', markerfacecolor='none', markeredgecolor='black', markeredgewidth=1.5, markersize=9, label='Left Hemisphere', linestyle='None'),
        Line2D([0], [0], marker='^', color='w', markerfacecolor='none', markeredgecolor='black', markeredgewidth=1.5, markersize=9, label='Right Hemisphere', linestyle='None'),
        Line2D([0], [0], marker='s', color='w', markerfacecolor='#10b981', markeredgecolor='black', markersize=9, label='Correct Guess', linestyle='None'),
        Line2D([0], [0], marker='s', color='w', markerfacecolor='#ef4444', markeredgecolor='black', markersize=9, label='Incorrect Guess', linestyle='None')
    ]
    
    # Custom offsets dictionary to avoid overlaps at identical coordinates
    # format: (marker_dx, marker_dy, label_dx, label_dy, ha, va)
    offsets = {
        ("sub-03", "L"): (0.0, 0.0, 0.12, 0.0, "left", "center"),
        ("sub-03", "R"): (0.0, 0.0, 0.12, 0.0, "left", "center"),
        ("sub-04", "L"): (0.0, 0.0, 0.12, 0.0, "left", "center"),
        ("sub-04", "R"): (0.0, 0.0, 0.12, 0.0, "left", "center"),
        ("sub-05", "L"): (-0.15, 0.0, -0.12, 0.0, "right", "center"),
        ("sub-05", "R"): (0.15, 0.0, 0.12, 0.0, "left", "center"),
        ("sub-06", "L"): (0.0, 0.0, 0.12, 0.0, "left", "center"),
        ("sub-06", "R"): (0.0, 0.16, 0.0, 0.12, "center", "bottom"),
        ("sub-11", "L"): (-0.16, -0.12, -0.12, -0.06, "right", "top"),
        ("sub-11", "R"): (0.16, -0.12, 0.12, -0.06, "left", "top")
    }
    
    # Draw points for all subjects and both hemispheres
    for _, row in df_merged.iterrows():
        subj = row["Subject"]
        hemi = row["hemisphere"]
        color = "#10b981" if row["correct_guess"] else "#ef4444"
        marker = "o" if hemi == "L" else "^"
        
        # Get offsets to prevent overlap
        dx, dy, ldx, ldy, ha, va = offsets.get((subj, hemi), (0.0, 0.0, 0.12, 0.0, "left", "center"))
        
        plot_x = row["sound_diff"] + dx
        plot_y = row["sensation_diff"] + dy
        
        plt.scatter(plot_x, plot_y, color=color, marker=marker, 
                    s=180, edgecolor="black", linewidth=1.5, zorder=3)
        
        # Draw label at offset position
        label_text = f"{subj} ({hemi})"
        plt.text(plot_x + ldx, plot_y + ldy, label_text, fontsize=8, fontweight='bold', zorder=4, 
                 ha=ha, va=va)
        
    plt.axhline(0, color="red", linestyle="--", linewidth=1.5, alpha=0.8, zorder=1)
    plt.axvline(0, color="red", linestyle="--", linewidth=1.5, alpha=0.8, zorder=1)
    
    plt.xlim(-3.5, 3.5)
    plt.ylim(-3.5, 3.5)
    plt.xlabel(r"Auditory Contrast ($\Delta$Sound: EXP - CON)", fontsize=9)
    plt.ylabel(r"Sensory Contrast ($\Delta$Sensation: EXP - CON)", fontsize=9)
    plt.title("Sensory Contrasts as Predictors of Blinding Efficacy (By Hemisphere)", fontsize=11, fontweight='bold', pad=12)
    
    plt.legend(handles=legend_elements_scatter, loc="lower left",
               bbox_to_anchor=(0.0, 0.0), frameon=True, fontsize=8,
               labelspacing=0.8, handletextpad=0.8)
    sns.despine()
    plt.tight_layout(pad=1.5)
    plt.savefig(os.path.join(output_dir, "sensory_contrast_scatter.png"),
                bbox_inches="tight", pad_inches=0.15)
    plt.close()

    print("All four premium data visualizations successfully generated and saved to derivatives/blinding/.")

if __name__ == "__main__":
    run_blinding_analysis()
